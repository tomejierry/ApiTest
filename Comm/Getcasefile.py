from ApiTest.Comm.Filepath import casepath
from ApiTest.Comm.Logtype import loggings
from ApiTest.Config.case_cfg import *
from openpyxl.styles import PatternFill, colors
import openpyxl


class ReadExcel(object):
    def __init__(self, casecfg):
        self.dirname = casecfg['dirname']
        self.filename = casepath(self.dirname, casecfg['filename'])

        # self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(self.filename)  # 获取工作簿对象
        self.sh = self.wb[casecfg['sheet_name']]  # 选择表单
        self.start = casecfg['start']
        self.end = casecfg['end']

    def read_data(self):

        """读取数据"""

        datas = list(self.sh.rows)

        if len(datas) <= 1:
            return 'Use cases are incomplete'   # 用例文件内容不完整

        # 获取第一行的数据，作为字典的键值
        li1 = []
        for i in datas[0]:
            print(i.value)

            li1.append(i.value)

        # print(li1)

        # 创建一个空列表，用于存放到用例数据
        cases = []

        if type(self.start) != int and self.start is not None:
            loggings.info('start 参数错误')
            return 'Input strings are not supported'
        if type(self.end) != int and self.end is not None:
            loggings.info('end 参数错误')
            return 'Input strings are not supported'

        elif self.start is None or self.start <= 0:
            self.start = 1
            if self.end is None:
                pass
            elif self.end <= 1:
                loggings.info('获取用例默认从第一行开始')
                self.end = 2
        elif type(self.start) == int and self.end is None:
            pass
        elif self.start >= self.end:
            loggings.error('end 参数必须大于 start')
            return 'Please enter the correct range'
        # 遍历除第一行之外的数据
        for i in datas[self.start:self.end]:  # 切片
            li2 = []

            for c in i:  # 获取该行数据的值
                # values = c.value
                li2.append(c.value)  # 获取每个表格中的数据
            # cases.append(li2)
            # 将该行数据和第一行数据打包，打包转换为字典
            cases.append(dict(zip(li1, li2)))  # 与标题和表格中的数据分装成一个测试用例case,添加到空列表cases中

        return cases

    def write_data(self, actual, report_name):
        wb2 = openpyxl.Workbook()
        loggings.info('创建一个新的EXCEL')

        sh = wb2[wb2.sheetnames[0]]
        loggings.info('更改默认Sheet名称')
        sh.title = "接口用例"
        wb2.save(casepath(self.dirname, report_name))
        filename = casepath(self.dirname, report_name)
        wb2 = openpyxl.load_workbook(filename)
        sheet = wb2['接口用例']  # 获取工作簿对象

        max_row = self.sh.max_row  # 最大行数
        max_column = self.sh.max_column  # 最大列数
        i1 = '%s%d' % (chr(97 + max_column), 1)  # 实际结果行数坐标
        loggings.info('获取测试用例，复写到新的EXCEL')
        if type(self.start) != int and self.start is not None:
            loggings.info('start 参数错误')
            return 'Input strings are not supported'
        if type(self.end) != int and self.end is not None:
            loggings.info('end 参数错误')
            return 'Input strings are not supported'

        elif self.start is None or self.start <= 0:
            self.start = 1
            if self.end is None:
                self.end = max_row
            elif self.end <= 1:
                loggings.info('获取用例默认从第一行开始')
                self.end = 2
        elif type(self.start) == int and self.end is None:
            pass
        elif self.start >= self.end:
            loggings.error('end 参数必须大于 start')
            return 'Please enter the correct range'
        for m in range(self.start+1, self.end+1):
            for n in range(97, 97 + max_column):
                n = chr(n)

                i = '%s%d' % (n, 1)
                ii = '%s%d' % (n, m)

                cell1 = self.sh[i].value
                cell2 = self.sh[ii].value
                sheet[i].value = cell1   # 复制表头到单元格
                sheet[ii].value = cell2  # 复制用例内容到单元格

        loggings.info('判断表格中实际结果的插入位置列:' + chr(97 + max_column))

        red_fill = PatternFill('solid', fgColor='FF0033')      # 设置背景颜色：红色
        green_fill = PatternFill('solid', fgColor='008000')    # 设置背景颜色：绿色
        yellow_fill = PatternFill('solid', fgColor='FFFF00')   # 设置背景颜色：黄色
        start = 0

        for number in range(self.start + 1, self.end+1):
            for re in [actual[start]]:  # 循环读取实际结果re[2]为结果状态
                i10 = '%s%d' % (chr(97 + max_column), number)

                if re[2] == 'failed':  # 根据用例状态改变EXCEL单元格背景颜色
                    sheet[i10].fill = red_fill
                elif re[2] == 'passed':
                    sheet[i10].fill = green_fill
                elif re[2] == 'error':
                    sheet[i10].fill = yellow_fill
                sheet[i10].value = re[2]  # 赋值给最后一列

                start += 1
        loggings.info('把执行结果插入到EXCEL中，生成测试结果报告保存在--' + casepath(self.dirname, report_name))
        sheet[i1].value = '实际结果'
        wb2.save(casepath(self.dirname, report_name))  # 保存数据
        wb2.close()


if __name__ == '__main__':
    data = ReadExcel(bxtAPP_casecfg())
    # data.write_data([['失败'], ['异常'], ['通过'], ['失败'], ['失败'], ['通过'], ['结果7'], ['结果8'], ['结果X']])

    case1 = data.read_data()
    print(case1)
